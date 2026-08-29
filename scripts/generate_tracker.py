#!/usr/bin/env python3
"""
90-Day AI Platform Action Tracker for Kennedy Hachimeganum
"""
import os
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

OUTPUT_DIR = '/home/z/my-project/download'
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'Kennedy_90Day_Tracker.xlsx')

wb = Workbook()

# ─── Colors ───
HEADER_FILL = PatternFill(start_color='2F454F', end_color='2F454F', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
ACCENT_FILL = PatternFill(start_color='1F6C92', end_color='1F6C92', fill_type='solid')
ACCENT_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
SUCCESS_FILL = PatternFill(start_color='427E56', end_color='427E56', fill_type='solid')
WARNING_FILL = PatternFill(start_color='978051', end_color='978051', fill_type='solid')
ERROR_FILL = PatternFill(start_color='AE4D44', end_color='AE4D44', fill_type='solid')
STRIPE_FILL = PatternFill(start_color='F0F2F2', end_color='F0F2F2', fill_type='solid')
LIGHT_FILL = PatternFill(start_color='EAECEE', end_color='EAECEE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')

BODY_FONT = Font(name='Calibri', size=10)
BODY_FONT_BOLD = Font(name='Calibri', size=10, bold=True)
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='2F454F')
SUBTITLE_FONT = Font(name='Calibri', size=11, bold=True, color='1F6C92')
LINK_FONT = Font(name='Calibri', size=10, color='1F6C92', underline='single')
MUTED_FONT = Font(name='Calibri', size=9, color='767D80')

THIN_BORDER = Border(
    left=Side(style='thin', color='BCCCD4'),
    right=Side(style='thin', color='BCCCD4'),
    top=Side(style='thin', color='BCCCD4'),
    bottom=Side(style='thin', color='BCCCD4')
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header_row(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row, cols, stripe=False):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.alignment = LEFT_WRAP
        cell.border = THIN_BORDER
        if stripe:
            cell.fill = STRIPE_FILL


# ════════════════════════════════════════════════════════════════
# SHEET 1: Platform Registration Tracker
# ════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Registration Tracker'

ws1.cell(row=1, column=1, value='AI Platform Registration & Status Tracker').font = TITLE_FONT
ws1.cell(row=2, column=1, value='Wike-Young Kennedy Hachimeganum | Port Harcourt, Nigeria').font = MUTED_FONT
ws1.cell(row=3, column=1, value='Track every platform: registration, verification, qualification, and first earnings.').font = MUTED_FONT

headers = ['Platform', 'Website', 'Priority', 'Registered?', 'Date Registered',
           'Profile Complete?', 'ID Verified?', 'PayPal Connected?',
           'Qualification Test', 'Test Score/Result', 'First Task Date',
           'First Earning ($)', 'Status', 'Notes']
for col, h in enumerate(headers, 1):
    ws1.cell(row=5, column=col, value=h)
style_header_row(ws1, 5, len(headers))

platforms = [
    ['CrowdGen (Appen)', 'crowdgen.com', 'TIER 1', '', '', '', '', '', '', '', '', '', 'Not Started', 'Register first. Highlight Ikwerre + English skills.'],
    ['TELUS Digital', 'telusinternational.ai', 'TIER 1', '', '', '', '', '', '', '', '', '', 'Not Started', 'Prepare for search evaluation assessment.'],
    ['Prolific', 'prolific.com', 'TIER 1', '', '', '', '', '', '', '', '', '', 'Not Started', 'Complete demographic profile fully for more studies.'],
    ['Alignerr', 'alignerr.com', 'TIER 2', '', '', '', '', '', '', '', '', '', 'Not Started', 'Complete AI interview with Zara. Test Nigeria access.'],
    ['Welocalize', 'welodata.ai', 'TIER 2', '', '', '', '', '', '', '', '', '', 'Not Started', 'Emphasize Ikwerre language for localization projects.'],
    ['Mercor', 'work.mercor.com', 'TIER 2', '', '', '', '', '', '', '', '', '', 'Not Started', 'Apply and test. If blocked, move on immediately.'],
    ['Outlier', 'outlier.ai', 'TIER 3', '', '', '', '', '', '', '', '', '', 'Not Started', 'May be geo-blocked. Try once, accept result.'],
]

for i, platform in enumerate(platforms):
    row = 6 + i
    for col, val in enumerate(platform, 1):
        ws1.cell(row=row, column=col, value=val)
    style_data_row(ws1, row, len(headers), stripe=(i % 2 == 0))

# Priority column coloring
for row in range(6, 6 + len(platforms)):
    cell = ws1.cell(row=row, column=3)
    if 'TIER 1' in str(cell.value):
        cell.fill = SUCCESS_FILL
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    elif 'TIER 2' in str(cell.value):
        cell.fill = WARNING_FILL
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    elif 'TIER 3' in str(cell.value):
        cell.fill = ERROR_FILL
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

# Column widths
col_widths = [20, 22, 10, 12, 14, 12, 12, 14, 14, 14, 14, 12, 12, 45]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Dropdown hints
ws1.cell(row=14, column=1, value='Dropdown Options:').font = BODY_FONT_BOLD
ws1.cell(row=15, column=1, value='Registered? = Yes / No / Blocked').font = MUTED_FONT
ws1.cell(row=16, column=1, value='Status = Not Started / Registered / Testing / Active / Earning / Blocked / Paused').font = MUTED_FONT


# ════════════════════════════════════════════════════════════════
# SHEET 2: Daily Activity Log
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Daily Activity Log')

ws2.cell(row=1, column=1, value='Daily Activity & Earnings Log').font = TITLE_FONT
ws2.cell(row=2, column=1, value='Log every work session. Track hours, tasks, quality scores, and earnings by platform.').font = MUTED_FONT

headers2 = ['Date', 'Day #', 'Platform', 'Hours Worked', 'Tasks Completed',
            'Task Type', 'Quality Score', 'Earnings ($)',
            'Cumulative ($)', 'Payment Received?', 'Notes / Issues']
for col, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=col, value=h)
style_header_row(ws2, 4, len(headers2))

# Pre-fill 90 days
from datetime import datetime, timedelta
start_date = datetime(2026, 7, 30)
for day in range(90):
    row = 5 + day
    date = start_date + timedelta(days=day)
    ws2.cell(row=row, column=1, value=date.strftime('%Y-%m-%d'))
    ws2.cell(row=row, column=2, value=day + 1)
    ws2.cell(row=row, column=9, value=0)  # Cumulative formula placeholder
    style_data_row(ws2, row, len(headers2), stripe=(day % 2 == 0))

# Cumulative formula (simple: adds earning to previous cumulative)
# We'll add formulas after the data
for day in range(90):
    row = 5 + day
    earning_cell = f'H{row}'
    cum_cell = f'I{row}'
    if day == 0:
        ws2.cell(row=row, column=9).value = f'=H{row}'
    else:
        prev_cum = f'I{row-1}'
        ws2.cell(row=row, column=9).value = f'=IF(H{row}="",{prev_cum},{prev_cum}+H{row})'

col_widths2 = [12, 7, 18, 11, 14, 22, 12, 11, 13, 14, 40]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Summary row
summary_row = 5 + 90 + 1
ws2.cell(row=summary_row, column=1, value='TOTALS').font = BODY_FONT_BOLD
ws2.cell(row=summary_row, column=4, value=f'=SUM(D5:D{5+89})').font = BODY_FONT_BOLD
ws2.cell(row=summary_row, column=8, value=f'=SUM(H5:H{5+89})').font = BODY_FONT_BOLD


# ════════════════════════════════════════════════════════════════
# SHEET 3: Weekly Milestones
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Weekly Milestones')

ws3.cell(row=1, column=1, value='90-Day Weekly Milestones').font = TITLE_FONT
ws3.cell(row=2, column=1, value='Check off each milestone as you complete it. Be honest.').font = MUTED_FONT

headers3 = ['Week', 'Phase', 'Milestone', 'Target Date', 'Completed?', 'Actual Date', 'Notes']
for col, h in enumerate(headers3, 1):
    ws3.cell(row=4, column=col, value=h)
style_header_row(ws3, 4, len(headers3))

milestones = [
    # Phase 1: Foundation (Weeks 1-4)
    ['1', 'Phase 1: Foundation', 'Create professional Gmail account (wikeyoung41@gmail.com)', '2026-08-05', '', '', ''],
    ['1', 'Phase 1: Foundation', 'Set up PayPal Nigeria account and complete verification', '2026-08-05', '', '', ''],
    ['1', 'Phase 1: Foundation', 'Create Payoneer account as backup payment method', '2026-08-05', '', '', ''],
    ['1', 'Phase 1: Foundation', 'Set up dedicated workspace with reliable internet', '2026-08-05', '', '', ''],
    ['1', 'Phase 1: Foundation', 'Register on CrowdGen (crowdgen.com) - complete full profile', '2026-08-05', '', '', 'Highlight Ikwerre + English'],
    ['1', 'Phase 1: Foundation', 'Register on TELUS Digital (telusinternational.ai)', '2026-08-05', '', '', ''],
    ['1', 'Phase 1: Foundation', 'Register on Prolific (prolific.com) - complete demographic profile', '2026-08-05', '', '', 'Fill in ALL details'],
    ['2', 'Phase 1: Foundation', 'Register on Alignerr (alignerr.com) - complete AI interview', '2026-08-12', '', '', 'Test Nigeria availability'],
    ['2', 'Phase 1: Foundation', 'Register on Welocalize/Welo Data (welodata.ai)', '2026-08-12', '', '', 'Emphasize Ikwerre language'],
    ['2', 'Phase 1: Foundation', 'Register on Mercor (work.mercor.com) - test access', '2026-08-12', '', '', 'If blocked, move on'],
    ['2', 'Phase 1: Foundation', 'Attempt Outlier registration (outlier.ai)', '2026-08-12', '', '', 'May be geo-blocked'],
    ['2', 'Phase 1: Foundation', 'Take typing speed test (target: 30+ WPM)', '2026-08-12', '', '', 'Use 10fastfingers.com'],
    ['3', 'Phase 1: Foundation', 'Begin CrowdGen qualification tests', '2026-08-19', '', '', 'Study guidelines first!'],
    ['3', 'Phase 1: Foundation', 'Complete TELUS Digital assessments', '2026-08-19', '', '', 'Search evaluation practice'],
    ['3', 'Phase 1: Foundation', 'Start accepting Prolific studies', '2026-08-19', '', '', 'Build track record'],
    ['3', 'Phase 1: Foundation', 'Take Welocalize language proficiency tests', '2026-08-19', '', '', 'English + Ikwerre'],
    ['4', 'Phase 1: Foundation', 'Complete all pending qualification tests', '2026-08-26', '', '', ''],
    ['4', 'Phase 1: Foundation', 'Earn first payment on ANY platform', '2026-08-26', '', '', 'Track in Daily Log'],
    ['4', 'Phase 1: Foundation', 'Review and update all platform profiles', '2026-08-26', '', '', ''],

    # Phase 2: Building Momentum (Weeks 5-8)
    ['5', 'Phase 2: Momentum', 'Achieve consistent 3-5 hour daily work schedule', '2026-09-02', '', '', ''],
    ['5', 'Phase 2: Momentum', 'Maintain quality score above 90% on all active platforms', '2026-09-02', '', '', ''],
    ['5', 'Phase 2: Momentum', 'Reach $20-$80 total earnings milestone', '2026-09-02', '', '', ''],
    ['6', 'Phase 2: Momentum', 'Apply for specialized projects on active platforms', '2026-09-09', '', '', 'Translation, localization'],
    ['6', 'Phase 2: Momentum', 'Create LinkedIn profile (linkedin.com)', '2026-09-09', '', '', 'Title: AI Data Annotator'],
    ['6', 'Phase 2: Momentum', 'Start basic prompt engineering learning (free course)', '2026-09-09', '', '', 'LearnPrompting.org or YouTube'],
    ['7', 'Phase 2: Momentum', 'Improve typing speed to 35+ WPM', '2026-09-16', '', '', 'Daily practice'],
    ['7', 'Phase 2: Momentum', 'Reach $100-$150 total earnings', '2026-09-16', '', '', ''],
    ['7', 'Phase 2: Momentum', 'Network with Nigerian contributors on Reddit', '2026-09-16', '', '', 'r/WFHJobs, r/WorkOnline'],
    ['8', 'Phase 2: Momentum', 'Drop platforms with no income (OneForma, Handshake)', '2026-09-23', '', '', 'Stop wasting time'],
    ['8', 'Phase 2: Momentum', 'Reach $200-$300 total earnings', '2026-09-23', '', '', ''],

    # Phase 3: Optimization (Weeks 9-13)
    ['9', 'Phase 3: Optimization', 'Focus on best-performing 2-3 platforms only', '2026-09-30', '', '', ''],
    ['9', 'Phase 3: Optimization', 'Apply for domain-specific projects (coding, translation)', '2026-09-30', '', '', 'If skills allow'],
    ['10', 'Phase 3: Optimization', 'Begin learning basic Python (freeCodeCamp)', '2026-10-07', '', '', 'Opens coding tasks'],
    ['10', 'Phase 3: Optimization', 'Update CV with actual AI training experience', '2026-10-07', '', '', ''],
    ['11', 'Phase 3: Optimization', 'Typing speed 40+ WPM', '2026-10-14', '', '', ''],
    ['11', 'Phase 3: Optimization', 'Reach $300-$450 total earnings', '2026-10-14', '', '', ''],
    ['12', 'Phase 3: Optimization', 'Explore direct freelance AI training opportunities', '2026-10-21', '', '', 'Beyond platforms'],
    ['12', 'Phase 3: Optimization', 'Set minimum hourly rate based on actual data', '2026-10-21', '', '', ''],
    ['13', 'Phase 3: Optimization', '90-DAY REVIEW: $300-$600 total earnings target', '2026-10-28', '', '', 'Typical: $400-$500'],
]

for i, ms in enumerate(milestones):
    row = 5 + i
    for col, val in enumerate(ms, 1):
        ws3.cell(row=row, column=col, value=val)
    style_data_row(ws3, row, len(headers3), stripe=(i % 2 == 0))
    # Phase coloring
    phase_cell = ws3.cell(row=row, column=2)
    if 'Phase 1' in str(phase_cell.value):
        phase_cell.fill = LIGHT_FILL
    elif 'Phase 2' in str(phase_cell.value):
        phase_cell.fill = YELLOW_FILL
    elif 'Phase 3' in str(phase_cell.value):
        phase_cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')

col_widths3 = [7, 22, 55, 14, 12, 14, 35]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════
# SHEET 4: Payment Tracker
# ════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Payment Tracker')

ws4.cell(row=1, column=1, value='Payment & Earnings Tracker').font = TITLE_FONT
ws4.cell(row=2, column=1, value='Track every payment received. Verify against platform dashboards.').font = MUTED_FONT

headers4 = ['Date', 'Platform', 'Amount ($)', 'Method', 'Reference/Transaction ID',
           'Received?', 'Date Received', 'Notes']
for col, h in enumerate(headers4, 1):
    ws4.cell(row=4, column=col, value=h)
style_header_row(ws4, 4, len(headers4))

# Pre-fill 30 rows
for row in range(5, 35):
    style_data_row(ws4, row, len(headers4), stripe=((row-5) % 2 == 0))

# Total row
total_row = 35
ws4.cell(row=total_row, column=1, value='TOTAL').font = BODY_FONT_BOLD
ws4.cell(row=total_row, column=3, value=f'=SUM(C5:C34)').font = Font(name='Calibri', size=11, bold=True, color='1F6C92')

col_widths4 = [14, 18, 12, 14, 25, 12, 14, 35]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════
# SHEET 5: Quick Reference
# ════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Quick Reference')

ws5.cell(row=1, column=1, value='Quick Reference Card').font = TITLE_FONT
ws5.cell(row=2, column=1, value='Keep this open while working. Key info at a glance.').font = MUTED_FONT

ref_data = [
    ['', '', ''],
    ['YOUR DETAILS', '', ''],
    ['Name', 'Wike-Young Kennedy Hachimeganum', ''],
    ['Email', 'wikeyoung41@gmail.com', ''],
    ['Phone', '+234 816 012 4516', ''],
    ['Location', 'Port Harcourt, Rivers State, Nigeria', ''],
    ['Languages', 'English (Fluent), Ikwerre (Native), Igbo (Partial)', ''],
    ['', '', ''],
    ['PAYMENT ACCOUNTS', '', ''],
    ['PayPal Email', 'wikeyoung41@gmail.com', 'Set up before first task'],
    ['Payoneer', '[Set up at payoneer.com]', 'Backup payment method'],
    ['', '', ''],
    ['PLATFORM LOGIN URLs', '', ''],
    ['CrowdGen', 'https://crowdgen.com', ''],
    ['TELUS Digital', 'https://www.telusinternational.ai', ''],
    ['Prolific', 'https://www.prolific.com/participants', ''],
    ['Alignerr', 'https://app.alignerr.com/signin', ''],
    ['Welocalize', 'https://welodata.ai/join-the-community', ''],
    ['Mercor', 'https://work.mercor.com/login', ''],
    ['Outlier', 'https://outlier.ai', ''],
    ['', '', ''],
    ['DAILY TARGETS', '', ''],
    ['Minimum Work Hours', '3-5 hours/day', ''],
    ['Target Quality Score', '90%+', ''],
    ['Target Typing Speed', '40 WPM by Day 90', ''],
    ['Earnings Target (Day 30)', '$20-$80', 'Conservative'],
    ['Earnings Target (Day 60)', '$100-$300', 'Conservative'],
    ['Earnings Target (Day 90)', '$300-$600', 'Conservative / Typical'],
    ['', '', ''],
    ['CRITICAL RULES', '', ''],
    ['Rule 1', 'NEVER use ChatGPT/AI to complete tasks', 'Instant permanent ban'],
    ['Rule 2', 'NEVER use VPN to bypass geo-restrictions', 'Instant permanent ban'],
    ['Rule 3', 'ALWAYS read task guidelines before starting', 'Prevents rejected work'],
    ['Rule 4', 'NEVER share your account', 'Instant permanent ban'],
    ['Rule 5', 'ALWAYS maintain quality score above 90%', 'Prevents project loss'],
    ['Rule 6', 'If a platform has no tasks for 2 weeks, move on', 'Do not waste time waiting'],
    ['Rule 7', 'Log every session in Daily Activity Log', 'Track progress'],
    ['', '', ''],
    ['SKILLS TO LEARN', '', ''],
    ['Priority 1', 'Typing speed improvement (practice daily)', '10fastfingers.com'],
    ['Priority 2', 'Prompt engineering basics', 'LearnPrompting.org'],
    ['Priority 3', 'Basic Python (for coding tasks)', 'freeCodeCamp.org'],
    ['Priority 4', 'Advanced reading comprehension', 'Read technical articles daily'],
]

for i, (col1, col2, col3) in enumerate(ref_data):
    row = 4 + i
    ws5.cell(row=row, column=1, value=col1)
    ws5.cell(row=row, column=2, value=col2)
    ws5.cell(row=row, column=3, value=col3)

    # Style headers
    if col2 == '' and col3 == '' and col1 != '':
        continue
    if col1 in ['YOUR DETAILS', 'PAYMENT ACCOUNTS', 'PLATFORM LOGIN URLs', 'DAILY TARGETS',
                'CRITICAL RULES', 'SKILLS TO LEARN']:
        for c in range(1, 4):
            ws5.cell(row=row, column=c).font = SUBTITLE_FONT
    elif col1.startswith('Rule') or col1.startswith('Priority'):
        ws5.cell(row=row, column=1).font = BODY_FONT_BOLD
        ws5.cell(row=row, column=2).font = BODY_FONT
        ws5.cell(row=row, column=3).font = MUTED_FONT
    else:
        ws5.cell(row=row, column=1).font = BODY_FONT_BOLD
        ws5.cell(row=row, column=2).font = BODY_FONT
        ws5.cell(row=row, column=3).font = MUTED_FONT

# URLs as hyperlinks
url_rows = {
    14: 'https://crowdgen.com',
    15: 'https://www.telusinternational.ai',
    16: 'https://www.prolific.com/participants',
    17: 'https://app.alignerr.com/signin',
    18: 'https://welodata.ai/join-the-community',
    19: 'https://work.mercor.com/login',
    20: 'https://outlier.ai',
}
for r, url in url_rows.items():
    cell = ws5.cell(row=r, column=2)
    cell.font = LINK_FONT
    cell.hyperlink = url

col_widths5 = [22, 50, 35]
for i, w in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════
wb.save(OUTPUT_FILE)
print(f'Tracker saved to: {OUTPUT_FILE}')
